"""
Excel reader with table detection.

Detects multiple tables within a single Excel sheet.
"""

import pandas as pd
import openpyxl
from typing import List, Dict, Tuple, Optional, NamedTuple


class TableRegion:
    """Represents a detected table region in an Excel sheet."""
    
    def __init__(self, start_row: int, end_row: int, start_col: int, end_col: int, name: str = "", title_row: Optional[int] = None):
        self.start_row = start_row
        self.end_row = end_row
        self.start_col = start_col
        self.end_col = end_col
        self.name = name
        self.title_row = title_row  # Row with section title (not part of data)
    
    def __repr__(self):
        return (f"TableRegion(name='{self.name}', "
                f"rows={self.start_row}-{self.end_row}, "
                f"cols={self.start_col}-{self.end_col})")


class ExcelReader:
    """
    Read Excel files and detect table regions.
    
    Usage:
        reader = ExcelReader('data.xlsx')
        tables = reader.detect_tables('Sheet1')
        df = reader.read_table(tables[0])
    """
    
    def __init__(self, filepath: str):
        """Initialize reader with Excel file path."""
        self.filepath = filepath
        self.workbook = openpyxl.load_workbook(filepath, data_only=True)
    
    @property
    def sheet_names(self) -> List[str]:
        """Get list of sheet names."""
        return self.workbook.sheetnames
    
    def detect_key_value_tables(self, sheet_name: str) -> List[TableRegion]:
        """
        Detect key-value format tables (for assumptions/parameters).
        
        A key-value table has:
        - Column A: Labels/parameter names
        - Column B: Values
        - Column C: Units (optional)
        
        Tables are separated by:
        - Empty rows (2+)
        - Title rows (A has text, B is empty)
        
        Args:
            sheet_name: Name of sheet to analyze
        
        Returns:
            List of TableRegion objects
        """
        ws = self.workbook[sheet_name]
        tables = []
        
        current_title = None
        current_title_row = None
        current_start = None
        current_end = None
        
        for row_idx in range(1, ws.max_row + 1):
            col_a = ws.cell(row_idx, 1).value
            col_b = ws.cell(row_idx, 2).value
            
            # Check if this is a title row (A has text, B is empty)
            if col_a is not None and col_b is None:
                # Save previous table if exists
                if current_start is not None:
                    tables.append(TableRegion(
                        start_row=current_start,
                        end_row=current_end,
                        start_col=1,
                        end_col=3,
                        name=current_title or f"Table_{len(tables)+1}",
                        title_row=current_title_row
                    ))
                
                # Start new table
                current_title = str(col_a).strip()
                current_title_row = row_idx
                current_start = None
                current_end = None
            
            # Check if this is a data row (A and B both have values)
            elif col_a is not None and col_b is not None:
                if current_start is None:
                    current_start = row_idx
                current_end = row_idx
            
            # Empty row
            elif col_a is None and col_b is None:
                # If we have accumulated data, save it
                if current_start is not None and current_end is not None:
                    tables.append(TableRegion(
                        start_row=current_start,
                        end_row=current_end,
                        start_col=1,
                        end_col=3,
                        name=current_title or f"Table_{len(tables)+1}",
                        title_row=current_title_row
                    ))
                    current_start = None
                    current_end = None
        
        # Don't forget the last table
        if current_start is not None and current_end is not None:
            tables.append(TableRegion(
                start_row=current_start,
                end_row=current_end,
                start_col=1,
                end_col=3,
                name=current_title or f"Table_{len(tables)+1}",
                title_row=current_title_row
            ))
        
        return tables
    
    def detect_tables(self, sheet_name: str, min_rows: int = 2) -> List[TableRegion]:
        """
        Detect tables in a sheet by finding contiguous data regions.
        
        Args:
            sheet_name: Name of the sheet to analyze
            min_rows: Minimum number of rows to consider a valid table
        
        Returns:
            List of TableRegion objects
        """
        ws = self.workbook[sheet_name]
        tables = []
        
        # Get all cells with values
        data_cells = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            for col_idx, value in enumerate(row, 1):
                if value is not None:
                    data_cells.append((row_idx, col_idx))
        
        if not data_cells:
            return []
        
        # Find contiguous regions (simple heuristic: separated by 2+ empty rows)
        data_cells.sort()
        
        current_region = {
            'start_row': data_cells[0][0],
            'end_row': data_cells[0][0],
            'start_col': min(c for r, c in data_cells),
            'end_col': max(c for r, c in data_cells),
        }
        
        for i in range(1, len(data_cells)):
            prev_row = data_cells[i-1][0]
            curr_row = data_cells[i][0]
            
            # If gap is > 1 row, start new region
            if curr_row - prev_row > 1:
                # Save current region if it has enough rows
                if current_region['end_row'] - current_region['start_row'] + 1 >= min_rows:
                    # Try to find table name (first row, first column)
                    name = ws.cell(current_region['start_row'], current_region['start_col']).value
                    if name:
                        name = str(name).strip()
                    
                    tables.append(TableRegion(
                        start_row=current_region['start_row'],
                        end_row=current_region['end_row'],
                        start_col=current_region['start_col'],
                        end_col=current_region['end_col'],
                        name=name or f"Table_{len(tables)+1}"
                    ))
                
                # Start new region
                current_region = {
                    'start_row': curr_row,
                    'end_row': curr_row,
                    'start_col': current_region['start_col'],
                    'end_col': current_region['end_col'],
                }
            else:
                current_region['end_row'] = curr_row
        
        # Don't forget the last region
        if current_region['end_row'] - current_region['start_row'] + 1 >= min_rows:
            name = ws.cell(current_region['start_row'], current_region['start_col']).value
            if name:
                name = str(name).strip()
            
            tables.append(TableRegion(
                start_row=current_region['start_row'],
                end_row=current_region['end_row'],
                start_col=current_region['start_col'],
                end_col=current_region['end_col'],
                name=name or f"Table_{len(tables)+1}"
            ))
        
        return tables
    
    def detect_header_row(self, sheet_name: str, region: TableRegion) -> Optional[int]:
        """
        Auto-detect which row contains headers in a table region.

        Args:
            sheet_name: Sheet name
            region: TableRegion to analyze

        Returns:
            Row index of header row (1-indexed), or None if no headers detected

        Detection logic:
        - Headers are typically strings when data rows are numeric
        - Headers often have unique values (no duplicates)
        - First row is most common header location
        """
        ws = self.workbook[sheet_name]

        # Check first few rows for header candidates
        max_check_rows = min(3, region.end_row - region.start_row + 1)

        for offset in range(max_check_rows):
            row_idx = region.start_row + offset
            row_values = [ws.cell(row_idx, col_idx).value
                         for col_idx in range(region.start_col, region.end_col + 1)]

            # Check if this row looks like headers
            # 1. All values should be non-None
            if not all(v is not None for v in row_values):
                continue

            # 2. Check next row to see if it's data (numeric)
            if row_idx < region.end_row:
                next_row_values = [ws.cell(row_idx + 1, col_idx).value
                                  for col_idx in range(region.start_col, region.end_col + 1)]

                # Count numeric values in next row
                numeric_count = sum(1 for v in next_row_values
                                  if isinstance(v, (int, float)) and v is not None)

                # Count string values in current row
                string_count = sum(1 for v in row_values
                                 if isinstance(v, str))

                # If current row is mostly strings and next row has numbers, this is likely header
                if string_count >= len(row_values) * 0.5 and numeric_count > 0:
                    return row_idx

            # If this is the first row and all strings, assume it's header
            if offset == 0 and all(isinstance(v, str) for v in row_values):
                return row_idx

        # Default: assume first row is header if it exists
        return region.start_row if region.end_row > region.start_row else None

    def read_table(self, region: TableRegion, has_header: bool = True, auto_detect_header: bool = True) -> pd.DataFrame:
        """
        Read a table region into a DataFrame.

        Args:
            region: TableRegion object
            has_header: If True, use headers (ignored if auto_detect_header=True)
            auto_detect_header: If True, automatically detect header row

        Returns:
            pandas DataFrame
        """
        ws = self.workbook.active

        # Extract data from region
        data = []
        for row_idx in range(region.start_row, region.end_row + 1):
            row_data = []
            for col_idx in range(region.start_col, region.end_col + 1):
                cell = ws.cell(row_idx, col_idx)
                row_data.append(cell.value)
            data.append(row_data)

        if not data:
            return pd.DataFrame()

        # Auto-detect header if enabled
        if auto_detect_header and len(data) > 1:
            # Check if first row looks like headers
            first_row = data[0]
            second_row = data[1] if len(data) > 1 else []

            # Count string vs numeric in first row
            first_strings = sum(1 for v in first_row if isinstance(v, str))
            # Count numeric in second row
            second_numeric = sum(1 for v in second_row if isinstance(v, (int, float)) and v is not None)

            # If first row is mostly strings and second has numbers, use first as header
            if first_strings >= len(first_row) * 0.5 and second_numeric > 0:
                df = pd.DataFrame(data[1:], columns=data[0])
            else:
                # No clear header detected
                df = pd.DataFrame(data)
        elif has_header and len(data) > 1:
            # First row as header
            df = pd.DataFrame(data[1:], columns=data[0])
        else:
            df = pd.DataFrame(data)

        return df
    
    def read_key_value_region(self, sheet_name: str, region: TableRegion) -> pd.DataFrame:
        """
        Read a key-value table region.
        
        Args:
            sheet_name: Sheet name
            region: TableRegion object
        
        Returns:
            DataFrame with first column as index
        """
        ws = self.workbook[sheet_name]
        
        # Extract data from region
        data = []
        for row_idx in range(region.start_row, region.end_row + 1):
            row_data = []
            for col_idx in range(region.start_col, region.end_col + 1):
                cell = ws.cell(row_idx, col_idx)
                row_data.append(cell.value)
            data.append(row_data)
        
        if not data:
            return pd.DataFrame()
        
        # Convert to DataFrame with proper column names
        if len(data[0]) == 2:
            df = pd.DataFrame(data, columns=['Parameter', 'Value'])
        elif len(data[0]) == 3:
            df = pd.DataFrame(data, columns=['Parameter', 'Value', 'Unit'])
        else:
            df = pd.DataFrame(data)
        
        # Set first column as index
        if not df.empty and 'Parameter' in df.columns:
            df.set_index('Parameter', inplace=True)
            df.index.name = None
        
        return df
    
    def read_key_value_table(self, sheet_name: str, 
                            start_row: Optional[int] = None,
                            end_row: Optional[int] = None) -> pd.DataFrame:
        """
        Read a key-value table (common in assumptions sheets).
        
        Format:
            Column A: Parameter name (index)
            Column B: Value
            Column C: Unit (optional)
        
        Args:
            sheet_name: Name of sheet to read
            start_row: Starting row (1-indexed), None = auto-detect
            end_row: Ending row (1-indexed), None = auto-detect
        
        Returns:
            DataFrame with parameter names as index
        """
        ws = self.workbook[sheet_name]
        
        # Auto-detect if not specified
        if start_row is None:
            start_row = 1
        if end_row is None:
            end_row = ws.max_row
        
        # Read data
        data = []
        for row_idx in range(start_row, end_row + 1):
            row = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
            # Skip completely empty rows
            if any(cell is not None for cell in row):
                data.append(row)
        
        if not data:
            return pd.DataFrame()
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Clean up: remove rows where first column is None
        df = df[df.iloc[:, 0].notna()].copy()
        
        # Set first column as index
        if not df.empty:
            df.set_index(df.columns[0], inplace=True)
            df.index.name = None
        
        return df
    
    def read_all_tables(self, sheet_name: str, style: str = "assumptions", key_value: bool = True) -> List:
        """
        Detect and read all tables from a sheet, returning styled tables.
        
        Args:
            sheet_name: Name of sheet to read
            style: Style preset to apply
            key_value: If True, use key-value table detection
        
        Returns:
            List of tuples (table_name, FinancialTable object)
        """
        from .styles import FinancialTable
        
        if key_value:
            regions = self.detect_key_value_tables(sheet_name)
        else:
            regions = self.detect_tables(sheet_name)
        
        tables = []
        
        for region in regions:
            if key_value:
                df = self.read_key_value_region(sheet_name, region)
            else:
                df = self.read_table(region)
            
            table = FinancialTable(df, style=style)
            tables.append((region.name, table))
        
        return tables


def list_sheets(filepath: str) -> list:
    """
    List all sheet names in an Excel file.

    Args:
        filepath: Path to Excel file

    Returns:
        List of sheet names

    Example:
        >>> from finmodel import list_sheets
        >>> sheets = list_sheets('data.xlsx')
        >>> print(sheets)
        ['Sheet1', 'Sheet2', 'Data']
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    return sheet_names


def list_tables(filepath: str, glimpse: bool = False, max_rows: int = 3) -> Dict:
    """
    List all tables in an Excel file by sheet.

    Args:
        filepath: Path to Excel file
        glimpse: If True, show headers and first N rows for each table
        max_rows: Number of rows to show in glimpse (default: 3)

    Returns:
        Dictionary mapping sheet names to list of table info dicts

    Example:
        >>> from finmodel import list_tables
        >>> # Just list tables
        >>> tables = list_tables('data.xlsx')
        >>> for sheet, tables_list in tables.items():
        ...     print(f"{sheet}: {[t['name'] for t in tables_list]}")
        >>>
        >>> # Show glimpse with headers and first 3 rows
        >>> tables = list_tables('data.xlsx', glimpse=True)
        >>> # Or show more rows
        >>> tables = list_tables('data.xlsx', glimpse=True, max_rows=5)
    """
    reader = ExcelReader(filepath)
    result = {}

    for sheet_name in reader.sheet_names:
        # Try key-value table detection first (for assumptions/parameters)
        regions = reader.detect_key_value_tables(sheet_name)

        if not regions:
            # Fallback to general table detection
            regions = reader.detect_tables(sheet_name)

        tables_info = []
        for region in regions:
            table_info = {
                'name': region.name,
                'rows': f"{region.start_row}-{region.end_row}",
                'cols': f"{region.start_col}-{region.end_col}",
                'size': f"{region.end_row - region.start_row + 1} rows × {region.end_col - region.start_col + 1} cols"
            }

            if glimpse:
                # Read table data
                df = reader.read_key_value_region(sheet_name, region)

                # Add glimpse info
                table_info['columns'] = df.columns.tolist() if hasattr(df, 'columns') else []
                table_info['preview'] = df.head(max_rows)

            tables_info.append(table_info)

        result[sheet_name] = tables_info

    return result


def print_tables(filepath: str, glimpse: bool = False, max_rows: int = 3):
    """
    Print all tables in an Excel file in a readable format.

    Args:
        filepath: Path to Excel file
        glimpse: If True, show headers and first N rows for each table
        max_rows: Number of rows to show in glimpse (default: 3)

    Example:
        >>> from finmodel import print_tables
        >>> # Just list tables
        >>> print_tables('data.xlsx')
        >>>
        >>> # Show glimpse with headers and first 3 rows
        >>> print_tables('data.xlsx', glimpse=True)
    """
    tables = list_tables(filepath, glimpse=glimpse, max_rows=max_rows)

    print(f"\n📊 Tables in {filepath}\n")
    print("=" * 70)

    for sheet_name, tables_list in tables.items():
        print(f"\n📄 Sheet: '{sheet_name}'")
        print("-" * 70)

        if not tables_list:
            print("  (no tables detected)")
            continue

        for i, table_info in enumerate(tables_list, 1):
            print(f"\n  {i}. '{table_info['name']}'")
            print(f"     Location: Rows {table_info['rows']}, Cols {table_info['cols']}")
            print(f"     Size: {table_info['size']}")

            if glimpse and 'preview' in table_info:
                print(f"\n     Preview:")
                # Print with indentation
                preview_str = str(table_info['preview'])
                for line in preview_str.split('\n'):
                    print(f"     {line}")

    print("\n" + "=" * 70)


def glimpse_table(filepath: str, sheet: str, max_rows: int = 3):
    """
    Show preview of all tables in a specific sheet.

    Args:
        filepath: Path to Excel file
        sheet: Sheet name to preview
        max_rows: Number of rows to show in preview (default: 3)

    Example:
        >>> from finmodel import glimpse_table
        >>> glimpse_table('data.xlsx', 'Sheet1')
        >>> glimpse_table('data.xlsx', 'Sheet1', max_rows=5)
    """
    reader = ExcelReader(filepath)

    if sheet not in reader.sheet_names:
        print(f"\n❌ Sheet '{sheet}' not found!")
        print(f"\nAvailable sheets:")
        for s in reader.sheet_names:
            print(f"  - '{s}'")
        return

    # Try key-value table detection first
    regions = reader.detect_key_value_tables(sheet)

    if not regions:
        # Fallback to general table detection
        regions = reader.detect_tables(sheet)

    print(f"\n📄 Sheet: '{sheet}'")
    print("=" * 70)

    if not regions:
        print("  (no tables detected)")
        return

    for i, region in enumerate(regions, 1):
        df = reader.read_key_value_region(sheet, region)

        print(f"\n  {i}. '{region.name}'")
        print(f"     Location: Rows {region.start_row}-{region.end_row}, Cols {region.start_col}-{region.end_col}")
        print(f"     Size: {region.end_row - region.start_row + 1} rows × {region.end_col - region.start_col + 1} cols")
        print(f"\n     Preview:")

        # Print with indentation
        preview_str = str(df.head(max_rows))
        for line in preview_str.split('\n'):
            print(f"     {line}")

    print("\n" + "=" * 70)


def finmodel_output(data, columns: list = None, title: str = "Ausgabe") -> 'FinancialTable':
    """
    Format model output/results with financial model styling.
    
    Creates a yellow/orange styled table for displaying final results.
    Excel convention: Yellow cells = outputs/results.
    
    Args:
        data: Dict, DataFrame, list of tuples, or list of lists
        columns: Column names (optional). If not provided:
                 - For dict/list of tuples: ['Value']
                 - For DataFrame: uses existing columns
                 - For list of lists: ['Value'] or auto-generated
        title: Section title (default: "Ausgabe") - deprecated
        
        Note: Use Quarto cell directives for captions and labels:
            #| label: tbl-name
            #| tbl-cap: "Caption text"
    
    Returns:
        FinancialTable with outputs styling (yellow background)
    
    Examples:
        >>> # Simple dict (single value column)
        >>> results = {'Call Price': 5.24, 'Put Price': 2.85}
        >>> finmodel_output(results)
        
        >>> # Dict with multiple columns
        >>> results = {
        ...     'Call': [5.24, 0.68, 0.04],
        ...     'Put': [2.85, -0.32, 0.04]
        ... }
        >>> finmodel_output(results, columns=['Price', 'Delta', 'Gamma'])
        
        >>> # List of tuples (single value)
        >>> results = [('Call Price', 5.24), ('Put Price', 2.85)]
        >>> finmodel_output(results)
        
        >>> # List of lists (multiple columns)
        >>> results = [
        ...     ('Call', 5.24, 0.68, 0.04),
        ...     ('Put', 2.85, -0.32, 0.04)
        ... ]
        >>> finmodel_output(results, columns=['Option', 'Price', 'Delta', 'Gamma'])
        
        >>> # From DataFrame (keeps columns)
        >>> df = pd.DataFrame({
        ...     'Price': [5.24, 2.85],
        ...     'Delta': [0.68, -0.32]
        ... }, index=['Call', 'Put'])
        >>> finmodel_output(df)
    """
    from .styles import FinancialTable
    
    # Convert to DataFrame
    if isinstance(data, pd.DataFrame):
        df = data.copy()
        # Apply custom column names if provided
        if columns is not None:
            if len(columns) != len(df.columns):
                raise ValueError(f"columns length ({len(columns)}) must match DataFrame columns ({len(df.columns)})")
            df.columns = columns
    
    elif isinstance(data, dict):
        # Check if dict values are lists/arrays (multiple columns)
        first_val = next(iter(data.values()))
        if isinstance(first_val, (list, tuple, pd.Series)):
            # Multiple columns: {'Call': [5.24, 0.68], 'Put': [2.85, -0.32]}
            df = pd.DataFrame(data).T
            if columns is not None:
                if len(columns) != len(df.columns):
                    raise ValueError(f"columns length ({len(columns)}) must match data columns ({len(df.columns)})")
                df.columns = columns
            else:
                # Auto-generate column names
                df.columns = [f'Col_{i+1}' for i in range(len(df.columns))]
        else:
            # Single column: {'Call Price': 5.24, 'Put Price': 2.85}
            df = pd.DataFrame(list(data.items()), columns=['Parameter', 'Value'])
            df.set_index('Parameter', inplace=True)
            if columns is not None:
                if len(columns) != 1:
                    raise ValueError(f"columns length ({len(columns)}) must be 1 for single-value dict")
                df.columns = columns
    
    elif isinstance(data, list):
        if not data:
            raise ValueError("data list cannot be empty")
        
        # Check if list of tuples/lists
        first_item = data[0]
        if isinstance(first_item, (list, tuple)):
            # List of lists/tuples
            if len(first_item) == 2:
                # Two columns: [(label, value), ...]
                df = pd.DataFrame(data, columns=['Parameter', 'Value'])
                df.set_index('Parameter', inplace=True)
                if columns is not None:
                    if len(columns) != 1:
                        raise ValueError(f"columns length ({len(columns)}) must be 1 for 2-element tuples")
                    df.columns = columns
            else:
                # More than 2 columns: [(label, val1, val2, ...), ...]
                if columns is not None:
                    if len(columns) != len(first_item) - 1:
                        raise ValueError(f"columns length ({len(columns)}) must be {len(first_item)-1}")
                    df = pd.DataFrame([row[1:] for row in data], 
                                     columns=columns,
                                     index=[row[0] for row in data])
                else:
                    # Auto-generate column names
                    col_names = [f'Col_{i+1}' for i in range(len(first_item)-1)]
                    df = pd.DataFrame([row[1:] for row in data],
                                     columns=col_names,
                                     index=[row[0] for row in data])
        else:
            raise ValueError("List items must be tuples or lists")
    
    else:
        raise ValueError("data must be dict, list of tuples/lists, or DataFrame")
    
    # Create output table with results (yellow) styling
    table = FinancialTable(df, style='results')
    
    return table
